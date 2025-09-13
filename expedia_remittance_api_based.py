import requests
import re
import json
import demjson3  # Install with: pip install demjson3
#from bs4 import BeautifulSoup  # Install with: pip install beautifulsoup4
from bs4 import BeautifulSoup
import openpyxl
from datetime import date
import os
import pandas as pd 
from datetime import datetime


def remark_excel(remark,row,hotel_id):
    workbook_n = openpyxl.load_workbook(name)
    sheet_n =workbook_n.active
    sheet_n.cell(row = row,column = 1).value = hotel_id
    sheet_n.cell(row = row,column = 8).value = remark
    workbook_n.save(name)
    workbook_n.close()


def update_hotel_no(hotel_no):
    workbook_l = openpyxl.load_workbook('LoginInput.xlsx')
    sheet_l = workbook_l.active
    sheet_l.cell(row = 2,column = 5).value = hotel_no
    workbook_l.save('LoginInput.xlsx')
    workbook_l.close()
    
    
    
    

today = date.today()
name = "Expediareservation_bsed_api " + today.strftime("%B_%d_%Y") + ".xlsx"
No = 1
while os.path.isfile(name):
    name = name.split(".xlsx")[0].split("[")[0] + "[" +str(No) + "]" + ".xlsx"
    No += 1
print("For OTA Finance Team Only  ----- ")
print(name)



df = pd.DataFrame(columns = ['Hotel id','Date requested','Status','Tracking number','Payout ID','Date paid','Amount processed','Remark'])
df.to_excel(name,'Sheet1',index = False)
row = 2
col = 2


workbook_l = openpyxl.load_workbook('LoginInput.xlsx')
sheet_l = workbook_l.active
s_date = sheet_l.cell(row = 2,column = 3).value
e_date = sheet_l.cell(row = 2,column = 4).value
JSESSIONID = sheet_l.cell(row = 2,column = 1).value
epcsid = sheet_l.cell(row = 2,column = 2).value
t  = sheet_l.cell(row = 2, column = 5).value
t = int(t)

# Convert to datetime.date objects
start_date = datetime.strptime(s_date, "%m/%d/%Y").date()
end_date = datetime.strptime(e_date, "%m/%d/%Y").date()



workbook_h = openpyxl.load_workbook('hotelID.xlsx')
sheet_h = workbook_h.active
rowss = sheet_h.max_row
for i in range (t,(rowss + 1)):
    update_hotel_no(i)
    if sheet_h.cell(row = i+1 ,column = 1 ).value == None:
        break
    else:
        #print(f"hotel Id {i} : {sheet_h.cell(row = i+1 ,column = 1 ).value}")
    
        url = "https://apps.expediapartnercentral.com/lodging/accounting/statementsAndInvoices.html"
        params = {"htid": sheet_h.cell(row = i+1 ,column = 1 ).value, "tab": "statements"}
        headers = {
            "User-Agent": "Mozilla/5.0",
            "Referer": "https://apps.expediapartnercentral.com/"
        }
        cookies = {
            "JSESSIONID": JSESSIONID,
            "epcsid": epcsid,
            "EPCSession": "8f6aaf99fa6f8ec5fd2220dc641baa91",
            "EG_SESSIONTOKEN": "fnkMtwrXZFQjeNsfQgeA3mTbWzsdzCDad08LiIFvprY:GyWvy968BGyNhdF0yXQj3u_VF-jeO0IEvqr5ZLMrjindnh3n-NXsDZb12QuvJ14ZBCqO3-SWQ0JHGF7fkc00xw"
        }
        response = requests.get(url, headers=headers, cookies=cookies, params=params)
        html = response.text
        
        soup = BeautifulSoup(html, "html.parser")
        scripts = soup.find_all("script")
        #print(scripts)
        target_script = None
        for script in scripts:
            if script.string and "window.statementsAndInvoices" in script.string:
                target_script = script.string
                break
        if target_script != None:
            if not target_script:
                print("❌ JS object not found in any <script> tag.")
                with open("expedia_debug.html", "w", encoding="utf-8") as f:
                    f.write(html)
                exit()
            match = re.search(r'window\.statementsAndInvoices\s*=\s*({.*?});', target_script, re.DOTALL)
            if not match:
                print("❌ Could not extract JS object from <script> tag.")
                exit()
            
            js_object_str = match.group(1)
            try:
                data = demjson3.decode(js_object_str)
            except Exception as e:
                print("❌ JSON parse error while decoding JS object:", e)
                print("🪵 Raw string that caused error:\n", js_object_str[:500])
                exit()
            payment_list = data.get("statementsAndInvoicesPayload", {}).get("statements", {}).get("paymentList", [])
            
            if not payment_list:
                print("⚠️ No data found.")
                remark_excel("⚠️ No data found.",row,sheet_h.cell(row = i+1 ,column = 1 ).value)
                row = row + 1
                exit()
            check = 0
            list_no=len(payment_list)
            for idx, p in enumerate(payment_list, start=1):
                
                date = p.get('dateRequested', 'None')
                if date != 'None':
                    requested_date = datetime.strptime(date, "%Y-%m-%d").date()
                    #print(f"start date : {s_date} resuested date : {date} end date : {e_date} ")
                    if start_date <= requested_date <= end_date:
                        check = check + 1
                        print(f" Hotel ID : {sheet_h.cell(row = i+1 ,column = 1 ).value}    tracking number : {p.get('transmissionQueueID', '')}✅    Remaining TN : {int(list_no)-int(idx)}")
                        '''
                        print(f"Your reference number: {p.get('paymentReferenceNumber', '')}")
                        print(f"Date requested       : {p.get('dateRequested', '')}")
                        print(f"Status               : {p.get('invoiceStatus', '')}")
                        print(f"Tracking number      : {p.get('transmissionQueueID', '')}")
                        print(f"Payout ID            : {p.get('invoiceId', '')}")
                        print(f"Date paid            : {p.get('datePaid', '')}")
                       # print(f"Payment statement    : {p.get('paymentRequestFilePath', '')}")
                        print(f"Amount processed     : {p.get('amountProcessed', 0)}")
                        print("-" * 50)
                        '''
                        workbook_n = openpyxl.load_workbook(name)
                        sheet_n =workbook_n.active
                        sheet_n.cell(row = row,column = 1).value = sheet_h.cell(row = i+1 ,column = 1 ).value
                        sheet_n.cell(row = row,column = 2).value = p.get('dateRequested', '')
                        sheet_n.cell(row = row,column = 3).value = p.get('invoiceStatus', '')
                        sheet_n.cell(row = row,column = 4).value = p.get('transmissionQueueID', '')
                        sheet_n.cell(row = row,column = 5).value = p.get('invoiceId', '')
                        sheet_n.cell(row = row,column = 6).value = p.get('datePaid', '')
                        sheet_n.cell(row = row,column = 7).value = p.get('amountProcessed', 0)
                        workbook_n.save(name)
                        workbook_n.close()
                        row = row + 1
                    else:
                        print(f" Hotel ID : {sheet_h.cell(row = i+1 ,column = 1 ).value}    tracking number : {p.get('transmissionQueueID', '')}❌    Remaining TN : {int(list_no)-int(idx)}")
            if check == 0 :
                remark_excel("⚠️ No data found In date range.",row,sheet_h.cell(row = i+1 ,column = 1 ).value)
                row = row + 1
                print("⚠️ No data found In date range.")
                
        else:
            remark_excel("⚠️ Hotel is Not found.",row,sheet_h.cell(row = i+1 ,column = 1 ).value)
            row = row + 1
            print(f"hotel Id {i} : {sheet_h.cell(row = i+1 ,column = 1 ).value} Not Found..") 
